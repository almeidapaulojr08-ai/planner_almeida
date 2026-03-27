import openpyxl, json
from datetime import datetime

wb = openpyxl.load_workbook('C:/Users/Dell/OneDrive/Área de Trabalho/Codes/Base/Despesas Paulo.xlsx', data_only=True)

BANK_MAP_CONTA = {
    'Nubank': {'id': 'nubank_conta', 'bank': 'nubank'},
    'Sicredi': {'id': 'sicredi_conta', 'bank': 'sicredi'},
}
BANK_MAP_CARTAO = {
    'Nubank': {'id': 'card_nubank', 'bank': 'nubank'},
    'Sicredi': {'id': 'card_sicredi', 'bank': 'sicredi'},
}

# Category + Subcategory mapping: excel_cat -> (app_cat, app_subcat)
CAT_MAP = {
    'Alimentação':   ('Alimentação', 'Geral'),
    'Mercado':       ('Mercado', ''),
    'Saúde':         ('Saúde', 'Geral'),
    'Assinaturas':   ('Assinaturas', ''),
    'Lazer':         ('Lazer', 'Geral'),
    'Viagem':        ('Viagem', 'Geral'),
    'Carro':         ('Carro', 'Manutenção'),
    'Vestuário':     ('Vestuário', ''),
    'Educação':      ('Educação', 'Cursos'),
    'Games':         ('Games', ''),
    'Presente':      ('Presente', ''),
    'Combustivel':   ('Carro', 'Combustível'),
    'Estacionamento':('Carro', 'Estacionamento'),
    'Shoppe/Temu':   ('Outros', ''),
    'Internet/Luz':  ('Casa', ''),
    'Aluguel':       ('Casa', 'Aluguel'),
    'Consórcio':     ('Outros', ''),
    'Casamento':     ('Outros', ''),
    'Documentos':    ('Outros', ''),
    'Emprestimos':   ('Outros', ''),
    'Poupanca':      ('Outros', ''),
    'Utensilios ':   ('Casa', 'Manutenção'),
    '-':             ('Outros', ''),
}

def refine_subcat(cat_raw, desc):
    desc_lower = desc.lower()

    if cat_raw == 'Internet/Luz':
        if 'copel' in desc_lower or 'luz' in desc_lower:
            return ('Casa', 'Luz')
        elif 'nio' in desc_lower or 'internet' in desc_lower:
            return ('Casa', 'Internet')
        elif 'agua' in desc_lower or 'sanepar' in desc_lower:
            return ('Casa', 'Água')
        elif 'gas' in desc_lower:
            return ('Casa', 'Gás')
        return ('Casa', 'Internet')

    if cat_raw == 'Saúde':
        if any(x in desc_lower for x in ['farmacia', 'drogaria', 'nissei', 'droga']):
            return ('Saúde', 'Farmácia')
        elif any(x in desc_lower for x in ['academia', 'growth', 'fitness']):
            return ('Saúde', 'Academia/Fitness')
        elif any(x in desc_lower for x in ['medico', 'cardio', 'clinica', 'hospital', 'inc ']):
            return ('Saúde', 'Médico')
        elif 'exame' in desc_lower:
            return ('Saúde', 'Exames')
        elif 'suplemento' in desc_lower:
            return ('Saúde', 'Suplementos')
        return ('Saúde', 'Geral')

    if cat_raw == 'Alimentação':
        if 'delivery' in desc_lower or 'ifood' in desc_lower:
            return ('Alimentação', 'Delivery')
        elif any(x in desc_lower for x in ['mercado', 'mercearia']):
            return ('Alimentação', 'Mercado')
        elif any(x in desc_lower for x in ['lanchonete', 'sorvete', 'chiquinho', 'tapioca']):
            return ('Alimentação', 'Lanchonete/Bar')
        else:
            return ('Alimentação', 'Restaurante')

    if cat_raw == 'Assinaturas':
        if any(x in desc_lower for x in ['netflix', 'spotify', 'youtube', 'disney', 'hbo', 'prime']):
            return ('Assinaturas', 'Streaming')
        elif any(x in desc_lower for x in ['microsoft', 'canva', 'icloud', 'anthropic', 'cursor', 'ai mirror', 'chatgpt']):
            return ('Assinaturas', 'Software/App')
        elif any(x in desc_lower for x in ['game', 'xbox', 'playstation']):
            return ('Assinaturas', 'Games')
        return ('Assinaturas', 'Outros')

    if cat_raw == 'Viagem':
        if any(x in desc_lower for x in ['hotel', 'pousada', 'hosped']):
            return ('Viagem', 'Hospedagem')
        elif any(x in desc_lower for x in ['pedagio', 'nutag', 'tag', 'uber']):
            return ('Viagem', 'Transporte')
        elif any(x in desc_lower for x in ['sorvet', 'almoco', 'lanche', 'restaurante', 'bapka']):
            return ('Viagem', 'Alimentação')
        elif any(x in desc_lower for x in ['carreiro', 'passeio', 'beto']):
            return ('Viagem', 'Passeio')
        return ('Viagem', 'Geral')

    if cat_raw == 'Lazer':
        if 'barbeiro' in desc_lower:
            return ('Lazer', 'Hobbies')
        elif any(x in desc_lower for x in ['show', 'evento', 'ingresso']):
            return ('Lazer', 'Shows/Eventos')
        return ('Lazer', 'Passeio')

    return None


transactions = []
idx = 1

ws = wb['Diario Despesas']
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    banco_raw, ref, data, desc, parcela, valor, metodo, cat_raw = row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]
    if not banco_raw or not data or valor is None:
        continue

    banco = str(banco_raw).strip()
    metodo_str = str(metodo).strip() if metodo else 'Débito'
    is_credito = 'dito' in metodo_str.lower() and 'cr' in metodo_str.lower()
    cat = str(cat_raw).strip() if cat_raw else '-'
    desc_str = str(desc).strip() if desc else 'Sem descrição'

    refined = refine_subcat(cat, desc_str)
    if refined:
        mapped_cat, mapped_subcat = refined
    else:
        mapped_cat, mapped_subcat = CAT_MAP.get(cat, ('Outros', ''))

    if is_credito:
        acc = BANK_MAP_CARTAO.get(banco, {'id': 'unknown', 'bank': 'unknown'})
    else:
        acc = BANK_MAP_CONTA.get(banco, {'id': 'unknown', 'bank': 'unknown'})

    if isinstance(data, datetime):
        date_str = data.strftime('%Y-%m-%d')
    else:
        date_str = str(data)[:10]

    raw_amount = float(valor)
    amount = abs(raw_amount)
    if amount == 0:
        continue

    parcela_str = str(parcela).strip().strip('()') if parcela else ''
    if parcela_str:
        desc_str += f' ({parcela_str})'

    # faturaRef from Referencia column (the billing month for credit card transactions)
    fatura_ref = ''
    if ref and hasattr(ref, 'strftime'):
        fatura_ref = ref.strftime('%Y-%m')

    # Auto-detect custoTipo
    CATEGORIAS_FIXAS = ['Assinaturas', 'Casa', 'Mercado', 'Saúde']
    custo_tipo = 'fixo' if mapped_cat in CATEGORIAS_FIXAS else 'variavel'

    t = {
        'id': f'imp_d_{idx}',
        'type': 'despesa',
        'desc': desc_str,
        'amount': round(amount, 2),
        'isNegative': raw_amount < 0,
        'category': mapped_cat,
        'subcategory': mapped_subcat,
        'date': date_str,
        'user': 'Paulo',
        'accountId': acc['id'],
        'formaPgto': 'credito' if is_credito else 'debito',
        'pago': True,
        'custoTipo': custo_tipo,
    }
    if fatura_ref:
        t['faturaRef'] = fatura_ref

    # Parcela info
    if parcela_str:
        t['parcela'] = parcela_str.strip('()')
        try:
            total_p = int(parcela_str.strip('()').split('/')[1])
            t['parcelaTotal'] = total_p
        except:
            pass
    transactions.append(t)
    idx += 1

# Entradas e Investimentos
ws2 = wb['Diario Entradas e Investimentos']
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=True):
    banco_raw, ref, data, desc, valor, tipo_raw = row[0], row[1], row[2], row[3], row[4], row[5]
    if not banco_raw or not data or valor is None:
        continue

    banco = str(banco_raw).strip()
    tipo = str(tipo_raw).strip().lower() if tipo_raw else 'receita'
    acc = BANK_MAP_CONTA.get(banco, {'id': 'unknown', 'bank': 'unknown'})

    if isinstance(data, datetime):
        date_str = data.strftime('%Y-%m-%d')
    else:
        date_str = str(data)[:10]

    raw_amount = float(valor)
    amount = abs(raw_amount)
    if amount == 0:
        continue

    desc_str = str(desc).strip() if desc else 'Sem descrição'

    if 'investimento' in tipo:
        tx_type = 'investimento'
        cat = 'Outros'
        if 'previd' in desc_str.lower():
            cat = 'Previdência'
        elif 'aplic' in desc_str.lower() or 'renda' in desc_str.lower():
            cat = 'Renda Fixa'
    else:
        tx_type = 'receita'
        cat = 'Outros'
        if 'sal' in desc_str.lower():
            cat = 'Salário'
        elif 'resg' in desc_str.lower():
            cat = 'Reembolso'

    t = {
        'id': f'imp_e_{idx}',
        'type': tx_type,
        'desc': desc_str,
        'amount': round(amount, 2),
        'isNegative': raw_amount < 0,
        'category': cat,
        'subcategory': '',
        'date': date_str,
        'user': 'Paulo',
        'accountId': acc['id'],
    }
    if tx_type == 'despesa':
        t['formaPgto'] = 'debito'
        t['pago'] = True

    transactions.append(t)
    idx += 1

# Stats
sub_counts = {}
for t in transactions:
    if t['type'] == 'despesa' and t.get('subcategory'):
        key = f"{t['category']} > {t['subcategory']}"
        sub_counts[key] = sub_counts.get(key, 0) + 1

print(f'Total: {len(transactions)}')
print('\nSubcategorias mapeadas:')
for k in sorted(sub_counts.keys()):
    print(f'  {k}: {sub_counts[k]}')

with open('C:/Users/Dell/financas-casal/import_data.js', 'w', encoding='utf-8') as f:
    f.write('const IMPORT_DATA = ')
    f.write(json.dumps(transactions, ensure_ascii=False, indent=2))
    f.write(';\n')

print('\nSalvo!')
